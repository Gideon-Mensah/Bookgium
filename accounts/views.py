from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib import messages
from django.urls import reverse_lazy
from django.db.models import Q, Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator
from django.core.exceptions import PermissionDenied
from django.template.loader import render_to_string
import mimetypes
import json
import csv
import io
from decimal import Decimal
from datetime import datetime, date, timedelta
from .models import Account, Transaction, JournalEntry, JournalEntryLine, SourceDocument
from .forms import (AccountForm, TransactionForm, AccountFilterForm, TransactionFilterForm,
                   JournalEntryForm, JournalEntryLineFormSet, QuickJournalEntryForm,
                   SourceDocumentForm, SourceDocumentFormSet, JournalEntrySourceDocumentFormSet)

# Helper function to check if user can access accounts
def can_access_accounts(user):
    """Check if user has permission to access accounting features"""
    return user.is_superuser or user.has_perm('accounts.view_account')

def calculate_period_opening_balance(account, from_date):
    """
    Calculate the opening balance for an account for a specific date period.
    This includes the account's opening balance plus all transactions before the from_date.
    """
    # Start with the account's opening balance
    opening_balance = account.opening_balance or Decimal('0')
    
    # Add effect of all transactions before the from_date
    pre_period_lines = JournalEntryLine.objects.filter(
        account=account,
        journal_entry__date__lt=from_date,
        journal_entry__is_posted=True
    ).aggregate(
        total_debits=Sum('amount', filter=Q(entry_type='debit')),
        total_credits=Sum('amount', filter=Q(entry_type='credit'))
    )
    
    # Calculate net effect of pre-period transactions
    total_debits = pre_period_lines['total_debits'] or Decimal('0')
    total_credits = pre_period_lines['total_credits'] or Decimal('0')
    
    # Calculate the final opening balance based on account type
    if account.account_type in ['asset', 'expense']:
        # For assets and expenses: debits increase balance, credits decrease balance
        net_pre_period_effect = total_debits - total_credits
        calculated_opening_balance = opening_balance + net_pre_period_effect
    else:
        # For liabilities, equity, and income: credits increase balance, debits decrease balance
        net_pre_period_effect = total_credits - total_debits
        calculated_opening_balance = opening_balance + net_pre_period_effect
    
    return calculated_opening_balance


# Account Views
@method_decorator(never_cache, name='dispatch')
class AccountListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Account
    template_name = 'accounts/account_list.html'
    context_object_name = 'accounts'
    paginate_by = 20

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_queryset(self):
        queryset = Account.objects.all().order_by('code')
        
        account_type = self.request.GET.get('account_type')
        is_active = self.request.GET.get('is_active')
        search = self.request.GET.get('search')

        if account_type:
            queryset = queryset.filter(account_type=account_type)
        
        if is_active == 'true':
            queryset = queryset.filter(is_active=True)
        elif is_active == 'false':
            queryset = queryset.filter(is_active=False)

        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(code__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = AccountFilterForm(self.request.GET)
        context['total_accounts'] = Account.objects.count()
        context['active_accounts'] = Account.objects.filter(is_active=True).count()
        return context

@method_decorator(never_cache, name='dispatch')
class AccountDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Account
    template_name = 'accounts/account_detail.html'
    context_object_name = 'account'

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account = self.get_object()
        
        # Get recent journal entry lines for this account
        recent_journal_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__is_posted=True
        ).select_related('journal_entry').order_by('-journal_entry__date', '-journal_entry__created_at')[:20]
        
        # Get all journal entry lines for balance calculation
        all_journal_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__is_posted=True
        ).select_related('journal_entry').order_by('-journal_entry__date', '-journal_entry__created_at')
        
        # Calculate running balance for each transaction
        running_balance = account.opening_balance
        transactions_with_balance = []
        
        # Reverse the order to calculate running balance correctly
        for line in reversed(list(all_journal_lines)):
            if account.account_type in ['asset', 'expense']:
                if line.entry_type == 'debit':
                    running_balance += line.amount
                else:
                    running_balance -= line.amount
            else:  # liability, equity, income
                if line.entry_type == 'credit':
                    running_balance += line.amount
                else:
                    running_balance -= line.amount
            
            line.running_balance = running_balance
            transactions_with_balance.append(line)
        
        # Reverse back to show most recent first and limit to recent transactions
        recent_transactions_with_balance = list(reversed(transactions_with_balance))[:20]
        
        # Calculate totals
        total_debits = all_journal_lines.filter(entry_type='debit').aggregate(
            total=Sum('amount'))['total'] or Decimal('0')
        total_credits = all_journal_lines.filter(entry_type='credit').aggregate(
            total=Sum('amount'))['total'] or Decimal('0')
        
        context.update({
            'recent_transactions': recent_transactions_with_balance,
            'total_debits': total_debits,
            'total_credits': total_credits,
            'transaction_count': all_journal_lines.count(),
        })
        
        return context

@method_decorator(never_cache, name='dispatch')
class AccountCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Account
    form_class = AccountForm
    template_name = 'accounts/account_form.html'
    success_url = reverse_lazy('accounts:account_list')

    def test_func(self):
        return can_access_accounts(self.request.user)

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, f'Account "{form.instance.name}" created successfully!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create New Account'
        context['submit_text'] = 'Create Account'
        return context

@method_decorator(never_cache, name='dispatch')
class AccountUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Account
    form_class = AccountForm
    template_name = 'accounts/account_form.html'

    def test_func(self):
        return can_access_accounts(self.request.user)

    def form_valid(self, form):
        messages.success(self.request, f'Account "{form.instance.name}" updated successfully!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Account: {self.object.name}'
        context['submit_text'] = 'Update Account'
        return context

@method_decorator(never_cache, name='dispatch')
class AccountDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Account
    template_name = 'accounts/account_confirm_delete.html'
    success_url = reverse_lazy('accounts:account_list')
    context_object_name = 'account'

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account = self.get_object()
        
        # Get transaction count and recent transactions
        transactions = account.transactions.all()
        context['transaction_count'] = transactions.count()
        context['recent_transactions'] = transactions.order_by('-date')[:5]
        
        return context

    def delete(self, request, *args, **kwargs):
        account = self.get_object()
        account_name = account.name
        
        # Check if account has transactions
        if account.transactions.exists():
            messages.error(request, f'Cannot delete account "{account_name}" because it has {account.transactions.count()} associated transactions.')
            return redirect('accounts:account_list')
        
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Account "{account_name}" deleted successfully!')
        return response

# Transaction Views
@method_decorator(never_cache, name='dispatch')
class TransactionListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Transaction
    template_name = 'accounts/transaction_list.html'
    context_object_name = 'transactions'
    paginate_by = 25

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_queryset(self):
        queryset = Transaction.objects.all().order_by('-date', '-created_at')
        
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        account_id = self.request.GET.get('account')
        search = self.request.GET.get('search')

        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(date__gte=date_from)
            except ValueError:
                pass

        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(date__lte=date_to)
            except ValueError:
                pass

        if account_id:
            queryset = queryset.filter(account_id=account_id)

        transaction_type = self.request.GET.get('transaction_type')
        amount_min = self.request.GET.get('amount_min')

        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)

        if amount_min:
            try:
                amount_min = float(amount_min)
                queryset = queryset.filter(amount__gte=amount_min)
            except ValueError:
                pass

        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(notes__icontains=search) |
                Q(account__name__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = TransactionFilterForm(self.request.GET)
        
        # Calculate summary statistics for filtered transactions
        queryset = self.get_queryset()
        credits = queryset.filter(transaction_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        debits = queryset.filter(transaction_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        
        context['total_credits'] = credits
        context['total_debits'] = debits
        context['net_balance'] = credits - debits
        
        return context

@method_decorator(never_cache, name='dispatch')
class TransactionDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Transaction
    template_name = 'accounts/transaction_detail.html'
    context_object_name = 'transaction'

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        transaction = self.get_object()
        
        # Get recent transactions for the same account
        context['recent_transactions'] = transaction.account.transactions.exclude(
            pk=transaction.pk
        ).order_by('-date')[:5]
        
        return context

@method_decorator(never_cache, name='dispatch')
class TransactionCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Transaction
    form_class = TransactionForm
    template_name = 'accounts/transaction_form.html'
    success_url = reverse_lazy('accounts:transaction_list')

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['document_formset'] = SourceDocumentFormSet(
                self.request.POST, 
                self.request.FILES, 
                instance=self.object
            )
        else:
            context['document_formset'] = SourceDocumentFormSet(instance=self.object)
        
        context['title'] = 'Create New Transaction'
        context['submit_text'] = 'Create Transaction'
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        document_formset = context['document_formset']
        
        if document_formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            
            document_formset.instance = self.object
            documents = document_formset.save(commit=False)
            for document in documents:
                document.uploaded_by = self.request.user
                document.save()
            
            # Handle deleted documents
            for document in document_formset.deleted_objects:
                document.delete()
            
            messages.success(self.request, 'Transaction created successfully!')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

@method_decorator(never_cache, name='dispatch')
class TransactionUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Transaction
    form_class = TransactionForm
    template_name = 'accounts/transaction_form.html'

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['document_formset'] = SourceDocumentFormSet(
                self.request.POST, 
                self.request.FILES, 
                instance=self.object
            )
        else:
            context['document_formset'] = SourceDocumentFormSet(instance=self.object)
        
        context['title'] = f'Edit Transaction: {self.object.description[:30]}'
        context['submit_text'] = 'Update Transaction'
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        document_formset = context['document_formset']
        
        if document_formset.is_valid():
            self.object = form.save()
            
            document_formset.instance = self.object
            documents = document_formset.save(commit=False)
            for document in documents:
                if not document.uploaded_by:
                    document.uploaded_by = self.request.user
                document.save()
            
            # Handle deleted documents
            for document in document_formset.deleted_objects:
                document.delete()
            
            messages.success(self.request, 'Transaction updated successfully!')
            return redirect(self.object.get_absolute_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))

@method_decorator(never_cache, name='dispatch')
class TransactionDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Transaction
    template_name = 'accounts/transaction_confirm_delete.html'
    success_url = reverse_lazy('accounts:transaction_list')
    context_object_name = 'transaction'

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        transaction = self.get_object()
        
        # Calculate what the balance would be after deletion
        current_balance = transaction.account.balance
        if transaction.transaction_type == 'credit':
            new_balance = current_balance - transaction.amount
        else:  # debit
            new_balance = current_balance + transaction.amount
            
        context['new_balance'] = new_balance
        return context

    def delete(self, request, *args, **kwargs):
        transaction = self.get_object()
        description = transaction.description[:50]
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Transaction "{description}" deleted successfully!')
        return response

# Dashboard and Report Views
@never_cache
@login_required
def accounts_dashboard(request):
    """Accounts dashboard with summary statistics"""
    from .utils import get_currency_symbol, get_user_currency
    
    if not can_access_accounts(request.user):
        messages.error(request, "You don't have permission to access accounts.")
        return redirect('dashboard')

    # Account summaries by type
    account_summary = {}
    for account_type, _ in Account.ACCOUNT_TYPES:
        accounts = Account.objects.filter(account_type=account_type, is_active=True)
        total_balance = sum(account.balance for account in accounts)
        account_summary[account_type] = {
            'count': accounts.count(),
            'total_balance': total_balance
        }

    # Recent journal entries
    recent_transactions = JournalEntry.objects.filter(is_posted=True).order_by('-date', '-created_at')[:10]

    # Monthly journal entry totals (last 6 months) - sum only debits to avoid double counting
    from datetime import datetime, timedelta
    six_months_ago = date.today() - timedelta(days=180)
    monthly_data = JournalEntry.objects.filter(
        date__gte=six_months_ago,
        is_posted=True
    ).annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        total=Sum('lines__amount', filter=Q(lines__entry_type='debit'))
    ).order_by('month')

    context = {
        'account_summary': account_summary,
        'recent_transactions': recent_transactions,
        'monthly_data': list(monthly_data),
        'total_accounts': Account.objects.count(),
        'active_accounts': Account.objects.filter(is_active=True).count(),
        'total_transactions': JournalEntry.objects.filter(is_posted=True).count(),
        'currency_symbol': get_currency_symbol(user=request.user),
        'currency_code': get_user_currency(request.user),
    }

    return render(request, 'accounts/dashboard.html', context)

@never_cache
@login_required  
def chart_of_accounts(request):
    """Chart of accounts view"""
    if not can_access_accounts(request.user):
        messages.error(request, "You don't have permission to access accounts.")
        return redirect('dashboard')

    accounts_by_type = {}
    type_totals = {}
    
    for account_type, type_name in Account.ACCOUNT_TYPES:
        accounts = Account.objects.filter(account_type=account_type, is_active=True).order_by('code')
        # Calculate type total
        type_total = sum(account.balance for account in accounts)
        accounts_by_type[type_name] = accounts
        type_totals[type_name] = type_total

    context = {
        'accounts_by_type': accounts_by_type,
        'type_totals': type_totals,
        'total_accounts': Account.objects.filter(is_active=True).count(),
    }

    return render(request, 'accounts/chart_of_accounts.html', context)

# Journal Entry Views (Double-Entry System)
@method_decorator(never_cache, name='dispatch')
class JournalEntryListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = JournalEntry
    template_name = 'accounts/journal_entry_list.html'
    context_object_name = 'journal_entries'
    paginate_by = 25

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_queryset(self):
        queryset = JournalEntry.objects.all().order_by('-date', '-created_at')
        
        # Filter by posting status
        is_posted = self.request.GET.get('is_posted')
        if is_posted == 'true':
            queryset = queryset.filter(is_posted=True)
        elif is_posted == 'false':
            queryset = queryset.filter(is_posted=False)
        
        # Filter by date range
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')
        
        if from_date:
            try:
                from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                queryset = queryset.filter(date__gte=from_date)
            except ValueError:
                pass
                
        if to_date:
            try:
                to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
                queryset = queryset.filter(date__lte=to_date)
            except ValueError:
                pass
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(reference__icontains=search) |
                Q(notes__icontains=search)
            )
        
        return queryset.select_related('created_by').prefetch_related('lines__account')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import get_currency_symbol
        context['currency_symbol'] = get_currency_symbol(user=self.request.user)
        
        # Add summary stats
        queryset = self.get_queryset()
        context['total_entries'] = queryset.count()
        context['posted_entries'] = queryset.filter(is_posted=True).count()
        context['unposted_entries'] = queryset.filter(is_posted=False).count()
        
        return context

@method_decorator(never_cache, name='dispatch')
class JournalEntryDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = JournalEntry
    template_name = 'accounts/journal_entry_detail.html'
    context_object_name = 'journal_entry'

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import get_currency_symbol
        context['currency_symbol'] = get_currency_symbol(user=self.request.user)
        
        # Add journal entry lines with proper ordering
        context['debit_lines'] = self.object.lines.filter(entry_type='debit').order_by('account__code')
        context['credit_lines'] = self.object.lines.filter(entry_type='credit').order_by('account__code')
        
        return context

@method_decorator(never_cache, name='dispatch')
class JournalEntryCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = JournalEntry
    template_name = 'accounts/journal_entry_form.html'
    fields = ['date', 'description', 'reference', 'notes']

    def test_func(self):
        return can_access_accounts(self.request.user)

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import get_currency_symbol
        context['currency_symbol'] = get_currency_symbol(user=self.request.user)
        context['accounts'] = Account.objects.filter(is_active=True).order_by('code')
        return context

@method_decorator(never_cache, name='dispatch')
class QuickJournalEntryCreateView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = 'accounts/quick_journal_entry_form.html'

    def test_func(self):
        return can_access_accounts(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import get_currency_symbol
        context['currency_symbol'] = get_currency_symbol(user=self.request.user)
        context['accounts'] = Account.objects.filter(is_active=True).order_by('code')
        return context

@never_cache
@login_required
def account_statement(request, account_id):
    """Detailed account statement with all transactions"""
    if not can_access_accounts(request.user):
        messages.error(request, "You don't have permission to access accounts.")
        return redirect('dashboard')
    
    account = get_object_or_404(Account, id=account_id)
    
    # Get date range filters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    export_format = request.GET.get('export')
    
    # Default date range (current year)
    if not from_date:
        from_date = date(date.today().year, 1, 1)
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        
    if not to_date:
        to_date = date.today()
    else:
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    
    # Calculate opening balance for the selected period
    # This includes all transactions before the from_date
    calculated_opening_balance = calculate_period_opening_balance(account, from_date)
    
    # Get all journal entry lines for this account in date range
    journal_lines = JournalEntryLine.objects.filter(
        account=account,
        journal_entry__date__gte=from_date,
        journal_entry__date__lte=to_date,
        journal_entry__is_posted=True
    ).select_related('journal_entry').order_by('journal_entry__date', 'journal_entry__id')
    
    # Calculate running balance starting from calculated opening balance
    transactions = []
    running_balance = calculated_opening_balance
    
    # Add opening balance entry if there are transactions or if opening balance is not zero
    if journal_lines.exists() or calculated_opening_balance != 0:
        # Determine how to display opening balance based on account type
        if account.account_type in ['asset', 'expense']:
            # Asset/Expense: Positive balance = Debit, Negative balance = Credit
            debit_opening = calculated_opening_balance if calculated_opening_balance > 0 else None
            credit_opening = abs(calculated_opening_balance) if calculated_opening_balance < 0 else None
        else:
            # Liability/Equity/Income: Positive balance = Credit, Negative balance = Debit
            credit_opening = calculated_opening_balance if calculated_opening_balance > 0 else None
            debit_opening = abs(calculated_opening_balance) if calculated_opening_balance < 0 else None
            
        transactions.append({
            'date': from_date,
            'description': 'Opening Balance',
            'reference': '',
            'debit': debit_opening,
            'credit': credit_opening,
            'balance': running_balance,
            'is_opening': True
        })
    
    # Process each transaction
    for line in journal_lines:
        if line.entry_type == 'debit':
            debit_amount = line.amount
            credit_amount = None
            # For Asset and Expense accounts: Debit increases balance
            # For Liability, Equity, and Income accounts: Debit decreases balance
            if account.account_type in ['asset', 'expense']:
                running_balance += line.amount
            else:  # liability, equity, income
                running_balance -= line.amount
        else:
            debit_amount = None
            credit_amount = line.amount
            # For Asset and Expense accounts: Credit decreases balance
            # For Liability, Equity, and Income accounts: Credit increases balance
            if account.account_type in ['asset', 'expense']:
                running_balance -= line.amount
            else:  # liability, equity, income
                running_balance += line.amount
            
        transactions.append({
            'date': line.journal_entry.date,
            'description': line.description or line.journal_entry.description,
            'reference': line.journal_entry.reference or f'JE-{line.journal_entry.id}',
            'debit': debit_amount,
            'credit': credit_amount,
            'balance': running_balance,
            'journal_entry_id': line.journal_entry.id,
            'is_opening': False
        })
    
    # Calculate totals
    total_debits = sum(t['debit'] for t in transactions if t['debit']) or Decimal('0')
    total_credits = sum(t['credit'] for t in transactions if t['credit']) or Decimal('0')
    closing_balance = running_balance
    
    context = {
        'account': account,
        'transactions': transactions,
        'from_date': from_date,
        'to_date': to_date,
        'opening_balance': calculated_opening_balance,
        'closing_balance': closing_balance,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'period_activity': total_debits - total_credits,
    }
    
    # Handle export requests
    if export_format == 'excel':
        return export_account_statement_excel(context)
    elif export_format == 'csv':
        return export_account_statement_csv(context)
    elif export_format == 'pdf':
        return export_account_statement_pdf(context)
    
    return render(request, 'accounts/account_statement.html', context)

def export_account_statement_excel(context):
    """Export account statement to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        # Return error response instead of redirect since this is called from another view
        response = HttpResponse('Excel export requires openpyxl. Please install it: pip install openpyxl', status=500)
        return response
    
    # Get organization settings
    from settings.models import CompanySettings
    try:
        company_settings = CompanySettings.objects.first()
        org_name = company_settings.organization_name if company_settings else "Your Organization Name"
    except:
        org_name = "Your Organization Name"
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Account Statement - {context['account'].code}"
    
    # Organization header
    ws['A1'] = org_name
    ws['A1'].font = Font(bold=True, size=18)
    ws['A2'] = "Account Statement"
    ws['A2'].font = Font(bold=True, size=16)
    
    # Account information
    ws['A4'] = f"Account: {context['account'].name}"
    ws['A4'].font = Font(bold=True)
    ws['A5'] = f"Account Code: {context['account'].code}"
    ws['A6'] = f"Account Type: {context['account'].get_account_type_display()}"
    ws['A7'] = f"Period: {context['from_date'].strftime('%Y-%m-%d')} to {context['to_date'].strftime('%Y-%m-%d')}"
    ws['A8'] = f"Generated: {date.today().strftime('%Y-%m-%d')}"
    
    # Add empty row
    row = 10
    
    # Column headers
    headers = ['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row += 1
    for transaction in context['transactions']:
        ws.cell(row=row, column=1, value=transaction['date'].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=transaction['description'])
        ws.cell(row=row, column=3, value=transaction['reference'])
        ws.cell(row=row, column=4, value=float(transaction['debit']) if transaction['debit'] else '')
        ws.cell(row=row, column=5, value=float(transaction['credit']) if transaction['credit'] else '')
        ws.cell(row=row, column=6, value=float(transaction['balance']))
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=2, value='Opening Balance:').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(context['opening_balance']))
    row += 1
    ws.cell(row=row, column=2, value='Total Debits:').font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(context['total_debits']))
    row += 1
    ws.cell(row=row, column=2, value='Total Credits:').font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(context['total_credits']))
    row += 1
    ws.cell(row=row, column=2, value='Closing Balance:').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(context['closing_balance']))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="account_statement_{context["account"].code}_{context["from_date"]}_{context["to_date"]}.xlsx"'
    
    wb.save(response)
    return response

def export_account_statement_csv(context):
    """Export account statement to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="account_statement_{context["account"].code}_{context["from_date"]}_{context["to_date"]}.csv"'
    
    # Get organization settings
    from settings.models import CompanySettings
    try:
        company_settings = CompanySettings.objects.first()
        org_name = company_settings.organization_name if company_settings else "Your Organization Name"
    except:
        org_name = "Your Organization Name"
    
    writer = csv.writer(response)
    
    # Organization and report header
    writer.writerow([org_name])
    writer.writerow(['Account Statement'])
    writer.writerow([])  # Empty row
    writer.writerow([f'Account: {context["account"].name}'])
    writer.writerow([f'Account Code: {context["account"].code}'])
    writer.writerow([f'Account Type: {context["account"].get_account_type_display()}'])
    writer.writerow([f'Period: {context["from_date"]} to {context["to_date"]}'])
    writer.writerow([f'Generated: {date.today()}'])
    writer.writerow([])  # Empty row
    
    # Column headers
    writer.writerow(['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Balance'])
    
    # Data rows
    for transaction in context['transactions']:
        writer.writerow([
            transaction['date'].strftime('%Y-%m-%d'),
            transaction['description'],
            transaction['reference'],
            transaction['debit'] if transaction['debit'] else '',
            transaction['credit'] if transaction['credit'] else '',
            transaction['balance']
        ])
    
    # Summary
    writer.writerow([])  # Empty row
    writer.writerow(['', 'Opening Balance:', '', '', '', context['opening_balance']])
    writer.writerow(['', 'Total Debits:', '', context['total_debits'], '', ''])
    writer.writerow(['', 'Total Credits:', '', '', context['total_credits'], ''])
    writer.writerow(['', 'Closing Balance:', '', '', '', context['closing_balance']])
    
    return response

def export_account_statement_pdf(context):
    """Export account statement to PDF format"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
    except ImportError:
        # Return error response instead of redirect since this is called from another view
        response = HttpResponse('PDF export requires reportlab. Please install it: pip install reportlab', status=500)
        return response
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Get organization settings
    from settings.models import CompanySettings
    try:
        company_settings = CompanySettings.objects.first()
        org_name = company_settings.organization_name if company_settings else "Your Organization Name"
    except:
        org_name = "Your Organization Name"
    
    # Organization name
    org_title = Paragraph(org_name, styles['Title'])
    elements.append(org_title)
    elements.append(Spacer(1, 6))
    
    # Report title
    title = Paragraph("Account Statement", styles['Heading1'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Account information
    account_info = [
        ['Account:', context['account'].name],
        ['Account Code:', context['account'].code],
        ['Account Type:', context['account'].get_account_type_display()],
        ['Period:', f"{context['from_date']} to {context['to_date']}"],
        ['Generated:', date.today().strftime('%Y-%m-%d')]
    ]
    
    info_table = Table(account_info, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Transaction table
    data = [['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Balance']]
    
    for transaction in context['transactions']:
        data.append([
            transaction['date'].strftime('%Y-%m-%d'),
            transaction['description'][:30] + '...' if len(transaction['description']) > 30 else transaction['description'],
            transaction['reference'],
            f"{transaction['debit']:.2f}" if transaction['debit'] else '',
            f"{transaction['credit']:.2f}" if transaction['credit'] else '',
            f"{transaction['balance']:.2f}"
        ])
    
    table = Table(data, colWidths=[1*inch, 2.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Description left-aligned
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Numbers right-aligned
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Summary table
    summary_data = [
        ['Opening Balance:', f"{context['opening_balance']:.2f}"],
        ['Total Debits:', f"{context['total_debits']:.2f}"],
        ['Total Credits:', f"{context['total_credits']:.2f}"],
        ['Closing Balance:', f"{context['closing_balance']:.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="account_statement_{context["account"].code}_{context["from_date"]}_{context["to_date"]}.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response

# Journal Entry Views
@login_required
def debug_revenue_account(request, account_id):
    """Debug function to investigate revenue account entries"""
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to access debug functions.")
        return redirect('dashboard')
    
    account = get_object_or_404(Account, id=account_id)
    
    # Get all journal entry lines for this account
    journal_lines = JournalEntryLine.objects.filter(
        account=account,
        journal_entry__is_posted=True
    ).select_related('journal_entry').order_by('journal_entry__date', 'journal_entry__id')
    
    # Analyze the entries
    debug_info = {
        'account': account,
        'total_entries': journal_lines.count(),
        'debit_entries': journal_lines.filter(entry_type='debit'),
        'credit_entries': journal_lines.filter(entry_type='credit'),
        'all_entries': journal_lines,
    }
    
    # Calculate totals
    debug_info['total_debits'] = debug_info['debit_entries'].aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    debug_info['total_credits'] = debug_info['credit_entries'].aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    debug_info['net_activity'] = debug_info['total_credits'] - debug_info['total_debits']
    
    return render(request, 'accounts/debug_revenue.html', debug_info)

@login_required
def revenue_accounts_overview(request):
    """Overview of all revenue accounts and their balances"""
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')
    
    # Get all income/revenue accounts
    revenue_accounts = Account.objects.filter(
        account_type='income',
        is_active=True
    ).order_by('code')
    
    # Add balance info to each account
    accounts_with_balances = []
    for account in revenue_accounts:
        account_info = {
            'account': account,
            'opening_balance': account.opening_balance,
            'journal_balance': account.journal_balance,
            'total_balance': account.balance,
            'has_negative_activity': account.journal_balance < 0,
            'transaction_count': account.journal_lines.filter(journal_entry__is_posted=True).count()
        }
        accounts_with_balances.append(account_info)
    
    context = {
        'accounts_with_balances': accounts_with_balances,
        'total_accounts': len(accounts_with_balances),
        'accounts_with_negative_activity': [acc for acc in accounts_with_balances if acc['has_negative_activity']],
    }
    
    return render(request, 'accounts/revenue_overview.html', context)

